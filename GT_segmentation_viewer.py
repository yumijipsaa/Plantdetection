# import some common libraries
import os, json, cv2, sys
# from google.colab.patches import cv2_imshow       # in colab
import numpy as np
# import some common detectron2 utilities
from detectron2.utils.visualizer import Visualizer
from detectron2.data import MetadataCatalog, DatasetCatalog

from openpyxl import Workbook
from openpyxl.drawing.image import Image



# if your dataset is in COCO format, this cell can be replaced by the following three lines:
# from detectron2.data.datasets import register_coco_instances
# register_coco_instances("my_dataset_train", {}, "json_annotation_train.json", "path/to/image/dir")
# register_coco_instances("my_dataset_val", {}, "json_annotation_val.json", "path/to/image/dir"
from detectron2.structures import BoxMode

from config import GT_imageConfig 
c_cfg = GT_imageConfig

dataset_path =  c_cfg.TRAIN_DIR
GT_image_path = c_cfg.GT_DIR_NAME
dataset_name = c_cfg.TRAIN_DATASET

# True이면 excel file에 image삽입. (용량 커짐)
just_view_gtimage =  c_cfg.JUST_VIEW_GTIMAGE
insert_image_excel = c_cfg.INSERT_IMAGE_EXCEL
excel_file_name = c_cfg.EXCEL_FILE_NAME

def get_dataset_dicts(img_dir, dataset_file):
    json_file = os.path.join(img_dir, dataset_file)

    with open(json_file) as f:
        imgs_anns = json.load(f)

    train_dataset_dicts_list = []

    c = 0
    for idx, metadata in enumerate(imgs_anns["metadatas"]):
        record = {}
        image_info = metadata["image_info"]
        file_name = image_info["file_name"]    # 2_20200825_093215.jpg
        filename = os.path.join(img_dir, file_name)     # training에 사용할 image의 path    

        record["file_name"] = filename                  # file의 path(name아님)
        record["image_id"] = image_info["image_id"]
        record["height"] = image_info["height"]
        record["width"] = image_info["width"]
        
        objs = []

        for _, instance_info in enumerate(metadata["instance_info"]):

            x_min, y_min, x_val, y_val = instance_info["bbox"]      # x_max = x_min + x_val (same y)
            obj = {
                "bbox":  [x_min, y_min, x_min + x_val, y_min + y_val],
                "bbox_mode": BoxMode.XYXY_ABS,
                "segmentation": instance_info["segmentation"],
                "category_id": instance_info["category_id"],
                }
            objs.append(obj)
            c +=1
        record["annotations"] = objs

        train_dataset_dicts_list.append(record)

    return train_dataset_dicts_list


def get_class_name(mood, dataset_file, c_cfg):
    json_file = os.path.join(c_cfg.TRAIN_DIR + "\\" + mood, dataset_file)

    with open(json_file) as f:
        imgs_anns = json.load(f)

    class_name = imgs_anns["classes"]

    return class_name


def compute_pay(image, d, class_name, ws, idx, file_name, image_number, total_pay, GT_dir_path):
    width, height = image.shape[0], image.shape[1]

    ws[f'A{idx}'] = "file_name"
    ws[f'B{idx}'] = file_name[-1]
    ws[f'C{idx}'] = f"number : {image_number}"
    idx += 1
    image_idx = idx

    count = 1
    pay = 0
    base_price = {'midrid': 52, 'leaf': 100, 'stem': 50, 'flower': 30, 'fruit': 100, 'cap': 52}
    for annotation in d["annotations"]:
        if class_name[annotation["category_id"]] == class_name[0]:          # midrid
            # print(f'{class_name[annotation["category_id"]]} : {110}')     
            pay += base_price[class_name[0]]
            ws[f'A{idx}'] = count
            ws[f'B{idx}'] = f'{class_name[annotation["category_id"]]}'
            ws[f'C{idx}'] = '110'
            idx +=1
            count +=1
        elif class_name[annotation["category_id"]] == class_name[1]:        # leaf
            # print(f'{class_name[annotation["category_id"]]} : {170}')
            pay += base_price[class_name[1]]
            ws[f'A{idx}'] = count
            ws[f'B{idx}'] = f'{class_name[annotation["category_id"]]}'
            ws[f'C{idx}'] = '170'
            idx +=1
            count +=1
        elif class_name[annotation["category_id"]] == class_name[2]:        # stem
            # print(f'{class_name[annotation["category_id"]]} : {80}')
            pay += base_price[class_name[2]]
            ws[f'A{idx}'] = count
            ws[f'B{idx}'] = f'{class_name[annotation["category_id"]]}'
            ws[f'C{idx}'] = '80'
            idx +=1
            count +=1
        elif class_name[annotation["category_id"]] == class_name[3]:        # flower
            # print(f'{class_name[annotation["category_id"]]} : {50}')
            pay += base_price[class_name[3]]
            ws[f'A{idx}'] = count
            ws[f'B{idx}'] = f'{class_name[annotation["category_id"]]}'
            ws[f'C{idx}'] = '50'
            idx +=1
            count +=1
        elif class_name[annotation["category_id"]] == class_name[4]:        # fruit
            # print(f'{class_name[annotation["category_id"]]} : {170}')
            pay += base_price[class_name[4]]
            ws[f'A{idx}'] = count
            ws[f'B{idx}'] = f'{class_name[annotation["category_id"]]}'
            ws[f'C{idx}'] = '170'
            idx +=1
            count +=1
        elif class_name[annotation["category_id"]] == class_name[5]:        # cap
            # print(f'{class_name[annotation["category_id"]]} : {110}')
            pay += base_price[class_name[5]]
            ws[f'A{idx}'] = count
            ws[f'B{idx}'] = f'{class_name[annotation["category_id"]]}'
            ws[f'C{idx}'] = '110'
            idx +=1
            count +=1

    ws[f'A{idx}'] = "pay"
    ws[f'B{idx}'] = pay
    if insert_image_excel:
        if count > 15:
            idx +=5
        elif count <= 15 and count > 10:
            idx +=10
        elif count <= 10 and count > 5:
            idx += 15
        else :
            idx += 20
    else :
        idx +=2

    total_pay +=pay

    # write text in image
    text = f"pay : {pay}"
    org = (int(width/10), int(height/10))
    fontFace = cv2.FONT_HERSHEY_SIMPLEX
    fontScale = 3
    thickness = 5
    color = (0, 0, 0)
    lineType = cv2.LINE_AA
    image = np.array(image)
    cv2.putText(image, text, org, fontFace, fontScale, color, thickness =thickness, lineType =lineType)

    # save image
    file_path = GT_dir_path + '\\' +file_name[-1]
    cv2.imwrite(file_path, image)

    if insert_image_excel:
        # excel에 image삽입
        img_excel = Image(file_path)
        if width > 2800 :
            img_excel.height = int(height/7)
            img_excel.width = int(width/7)
        elif width <= 2700 and width > 1600: 
            img_excel.height = int(height/4)
            img_excel.width = int(width/4)
        elif width <= 1600 and width > 1000:
            img_excel.height = int(height/2)
            img_excel.width = int(width/2)
        else :
            pass

        ws.add_image(img_excel, f'F{image_idx}')
     
    return image, ws, idx, total_pay

def GT_segmentation_viewer():
    DatasetCatalog.register(c_cfg.DATASET_NAME + "_" + "train", lambda d="train": get_dataset_dicts(c_cfg.TRAIN_DIR + "/train", c_cfg.TRAIN_DATASET))
    MetadataCatalog.get(c_cfg.DATASET_NAME + "_" + "train").set(thing_classes=get_class_name("train", c_cfg.TRAIN_DATASET, c_cfg))
    GT_metadata = MetadataCatalog.get(c_cfg.DATASET_NAME + "_" + "train")

    dataset_dicts = get_dataset_dicts(dataset_path + "/train", c_cfg.TRAIN_DATASET)

    class_name = get_class_name("train", c_cfg.TRAIN_DATASET, c_cfg)

    GT_dir_path = os.path.join(os.getcwd(), GT_image_path)
    os.makedirs(GT_dir_path, exist_ok = True)

    # excel instance 생성
    wb = Workbook()
    ws = wb.active

    idx = 1
    total_pay = 0
    for image_number, d in enumerate(dataset_dicts):
        print(f"-----({image_number + 1}/{len(dataset_dicts)}-----)")
        img = cv2.imread(d["file_name"])
        visualizer = Visualizer(img[:, :, ::-1], metadata=GT_metadata, scale=1)
        out = visualizer.draw_dataset_dict(d)
        image = out.get_image()[:, :, ::-1]

        file_name = d["file_name"].split('\\')

        image, ws, idx, total_pay = compute_pay(image, d, class_name, ws, idx, file_name, image_number+1, total_pay, GT_dir_path)
            
        #cv2.imshow("im", out.get_image()[:, :, ::-1])       # cv2_imshow(out.get_image()[:, :, ::-1])
        #while True:
        #    if cv2.waitKey() == 27:
        #        break
    
    if not just_view_gtimage:
        ws[f'D{idx}'] = "total pay"
        ws[f'E{idx}'] = total_pay
        # excel file로 저장
        wb.save(GT_dir_path + "/" + excel_file_name)



if __name__ == '__main__':
    GT_segmentation_viewer()


